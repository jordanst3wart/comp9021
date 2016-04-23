# Extracts from the list of countries accessible at
# http://www.worldbank.org/en/country
# population and GDP, when available,
# and creates a spreadsheet with those data.
#
# Written by Eric Martin for COMP9021

import os.path
import urllib.request
import sys

import bs4
import openpyxl


def convert_to_number(string):
    units = {'million': 10 ** 6, 'billion': 10 ** 9, 'trillion': 10 ** 12}
    # Most numbers are floating point numbers followed by one of the three units.
    for unit in units:
        if string.find(unit) >= 0:
            return int(float(string.rstrip().rstrip(unit)) * units[unit])
    # All other numbers use commas as thousands separators.
    return int(''.join(string.split(',')))      


def generate_country_population_and_gdp():
    # Example of html code being matched:
    #     <li class="name-country">
    #        <span><a href="http://www.worldbank.org/en/country/afghanistan">Afghanistan
    #              </a>
    #        </span>
    #     </li>
    for country in top_page.select('.name-country > span > a'):
        country_name = country.getText()
        try:
            with urllib.request.urlopen(country.get('href')) as country_url:
                country_page = bs4.BeautifulSoup(country_url, 'html.parser')
                # Example of html code being matched
                # (we are interested in only the first two, population and GDP, respectively):
                #     <td class="c01v1-country-amounts brd-t-0">31.63&#x9;million </td>
                #     <td class="c01v1-country-amounts">&#x24;20.04&#x9;billion </td>
                #     <td class="c01v1-country-amounts">1.3&#x25; </td>
                #     <td class="c01v1-country-amounts">4.6&#x25; </td>
                country_data = country_page.select('.c01v1-country-amounts')
                if len(country_data) < 2:
                    print('Could not access the desired data for {}.'.format(country_name))
                    continue
                country_population = country_data[0].getText()
                country_GDP = country_data[1].getText()
                if country_population.find('N/A') >= 0 or country_GDP.find('N/A') >= 0:
                    print('Could not access the desired data for {}.'.format(country_name))        
                    continue
                yield (country_name,
                       convert_to_number(country_population),
                       convert_to_number(country_GDP.lstrip().lstrip('$')))
        except urllib.error.HTTPError:
            print('Could not access the resource for {}.'.format(country_name))


file_name = 'country_population_and_gdp.xlsx'
if os.path.isfile(file_name):
    print('A file named {} already exists.'.format(file_name)) 
    print('You have to remove it first.')
    sys.exit()
try:
    with urllib.request.urlopen('http://www.worldbank.org/en/country') as top_url:
        top_page = bs4.BeautifulSoup(top_url, 'html.parser')
        workbook = openpyxl.Workbook()
        spreadsheet = workbook.active
        spreadsheet.title = 'World countries'
        spreadsheet['A1'] = 'Country'
        spreadsheet['B1'] = 'Population'
        spreadsheet['C1'] = 'GDP (in USD)'
        for counter, (country, population, GDP) in enumerate(generate_country_population_and_gdp(), 2):
            spreadsheet.cell(row = counter, column = 1).value = country
            spreadsheet.cell(row = counter, column = 2).value = population
            spreadsheet.cell(row = counter, column = 3).value = GDP
        workbook.save(file_name)
except urllib.error.HTTPError:
    print('Could not access the top resource.')
    sys.exit()
